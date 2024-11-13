from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
import pathlib

# Ckasse que representa a janela do sistema
class CadastroDeClientes:
    def __init__(self, root):

        # Criação da tela principal
        root.title("Cadastro de Clientes")
        mainframe = ttk.Frame(root, padding="3 3 12 12")
        mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
        mainframe.columnconfigure(0, weight=1)
        mainframe.rowconfigure(0, weight=1)

        # Frame que contém os botões
        frame_botoes = ttk.Frame(mainframe)
        frame_botoes.grid(column=3, row=3, sticky=(E, S))
        # Botão que vai salvar as mudanças
        botao_salvar = ttk.Button(frame_botoes, text='Salvar', command=self.salvar)
        botao_salvar.grid(column=1, row=0)

        # Botão que vai limpar os campos
        botao_limpar = ttk.Button(frame_botoes, text='Limpar', command=self.limpar)
        botao_limpar.grid(column=0, row=0)

        # Campo para escrever o nome completo
        frame_nome = ttk.Frame(mainframe)
        frame_nome.grid(column=0, row=0, columnspan=7, sticky=W)
        label_nome = ttk.Label(frame_nome, text='Nome Completo: ')
        label_nome.grid(column=0, row=0, sticky=W)
        self.nome = StringVar()
        entry_nome = ttk.Entry(frame_nome, width=61, textvariable=self.nome)
        entry_nome.grid(column=0, row=1, columnspan=7, sticky=W)

        # Campo para escrever o CPF
        frame_cpf = ttk.Frame(mainframe)
        frame_cpf.grid(column=0, row=1, sticky=W, columnspan=3)
        label_cpf = ttk.Label(frame_cpf, text='CPF:')
        label_cpf.grid(column=0, row=0, sticky=W)
        self.cpf = StringVar()
        entry_cpf = ttk.Entry(frame_cpf, width=30, textvariable=self.cpf)
        entry_cpf.grid(column=0, row=1, sticky=W, columnspan=3)

        # Campo para escrever o CEP
        frame_cep = ttk.Frame(mainframe)
        frame_cep.grid(column=3, row=1, sticky=W, columnspan=3)
        label_cep = ttk.Label(frame_cep, text='CEP:')
        label_cep.grid(column=0, row=0, sticky=W)
        self.cep = StringVar()
        entry_cep = ttk.Entry(frame_cep, width=25, textvariable=self.cep)
        entry_cep.grid(column=0, row=1, columnspan=3, sticky=W)

        # Campo para escrever o endereço
        frame_endereco = ttk.Frame(mainframe)
        frame_endereco.grid(column=0, row=2, sticky=W, columnspan=7)
        label_endereco = ttk.Label(frame_endereco, text='Endereço:')
        label_endereco.grid(column=0, row=0, sticky=W)
        self.endereco = StringVar()
        entry_endereco = ttk.Entry(frame_endereco, width=61, textvariable=self.endereco)
        entry_endereco.grid(column=0, row=1, columnspan=7, sticky=W)

        # Campo para escrever a data de nascimento
        frame_data_nasc = ttk.Frame(mainframe)
        frame_data_nasc.grid(column=0, row=3, sticky=W, columnspan=3)
        label_data_nasc = ttk.Label(frame_data_nasc, text='Data de Nascimento:')
        label_data_nasc.grid(column=0, row=0, sticky=W)
        self.data_nasc = StringVar()
        entry_data_nasc = ttk.Entry(frame_data_nasc, width=22, textvariable=self.data_nasc)
        entry_data_nasc.grid(column=0, row=1, columnspan=3)

        # Ajustes na exibição dos elementos
        for child in mainframe.winfo_children():
            child.grid_configure(padx=3, pady=3)

    # Função que salva os dados no arquivo
    def salvar(self, *args):
        # Verifica se o arquivo já existe. Caso não exista, ele o cria
        wb = pathlib.Path('Clientes.xlsx')
        if wb.exists():
            pass
        else:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Nome Completo'
            ws['B1'] = 'CPF'
            ws['C1'] = 'CEP'
            ws['D1'] = 'Endereço'
            ws['E1'] = 'Data de Nascimento'
            wb.save('Clientes.xlsx')
        # Captura e salvamento dos dados
        nome = self.nome.get()
        cpf = self.cpf.get()
        cep = self.cep.get()
        endereco = self.endereco.get()
        data_nasc = self.data_nasc.get()
        if nome == '' or cpf == '' or cep == '' or endereco == '' or data_nasc == '':
            messagebox.showerror("Sistema", "Erro! Todos os campos devem ser preenchidos!")
        else:
            wb = load_workbook('Clientes.xlsx')
            ws = wb.active
            ws.cell(column=1, row=ws.max_row+1, value=nome)
            ws.cell(column=2, row=ws.max_row, value=cpf)
            ws.cell(column=3, row=ws.max_row, value=cep)
            ws.cell(column=4, row=ws.max_row, value=endereco)
            ws.cell(column=5, row=ws.max_row, value=data_nasc)
            wb.save('Clientes.xlsx')
            messagebox.showinfo('Sistema', 'Dados salvos com sucesso!')

    # Função que limpa o conteúdo escrito nos campos
    def limpar(self, *args):
        self.nome.set('')
        self.cpf.set('')
        self.cep.set('')
        self.endereco.set('')
        self.data_nasc.set('')


root = Tk()
root.resizable(False, False)
CadastroDeClientes(root)
root.mainloop()

